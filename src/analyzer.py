import copy
import datetime
from pathlib import Path
import os

import cv2
import openpyxl
import scenedetect
from scenedetect import SceneManager, open_video, ContentDetector, FrameTimecode
import torch
from tqdm import tqdm
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from transformers import AutoProcessor, AutoModelForCausalLM, AutoConfig
from PIL import Image


ROOT_DIR = Path(os.path.abspath(__file__)).parent.parent.absolute()
DEFAULT_OUTPUT_DIR = os.path.join(ROOT_DIR, "output")
DEFAULT_MODEL = "MiaoshouAI/Florence-2-large-PromptGen-v2.0"
DEFAULT_WEIGHTS = ContentDetector.Components(delta_edges=.2)

device = "cuda:0" if torch.cuda.is_available() else "cpu"
torch_dtype = torch.float16 if torch.cuda.is_available() else torch.float32


class FrameAnalyzer():

    model = None
    processor = None

    def __init__(self, threshold=27, describe_frames=True, output_dir=DEFAULT_OUTPUT_DIR, model_name=DEFAULT_MODEL, weights=DEFAULT_WEIGHTS):
        self.describe_frames = describe_frames
        self.threshold = threshold
        self.weights = weights
        self.model_name = model_name
        self.img_scale = 33
        self.output_dir = output_dir

    def split_video_into_scenes(self, video, threshold=27.0, weights=ContentDetector.Components()):
        scene_manager = SceneManager()
        scene_manager.add_detector(ContentDetector(threshold=threshold, weights=weights))
        scene_manager.detect_scenes(video, show_progress=True) # , duration=FrameTimecode(2500, video.frame_rate)
        scene_list = scene_manager.get_scene_list()
        return scene_list
        
    def describe_frame(self, frame, max_tokens=1024):
        if self.model is None:
            self.model = AutoModelForCausalLM.from_pretrained(self.model_name, torch_dtype=torch_dtype, trust_remote_code=True).to(device)
        if self.processor is None:
            self.processor = AutoProcessor.from_pretrained(self.model_name, trust_remote_code=True)
        raw_image = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)).convert('RGB')
        inputs = self.processor(text="<DETAILED_CAPTION>", images=raw_image, return_tensors="pt").to(device, torch_dtype)
        outputs = self.model.generate(
            input_ids=inputs["input_ids"],
            pixel_values=inputs["pixel_values"],
            max_new_tokens=max_tokens,
            do_sample=False
        )
        description = self.processor.batch_decode(outputs, skip_special_tokens=True)[0]
        return description
    
    def analyze(self, video_paths):
        for vid_index, video_path in enumerate(video_paths):
            print("{d}: Starting Video {i}".format(i=vid_index+1, d=datetime.datetime.now().strftime("%Y_%m_%d_%H:%M:%S")))

            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            params_sheet = wb.create_sheet("Parameters")
            params_sheet.append(["Parameter", "Value"])
            params_sheet.append(["threshold", self.threshold])
            params_sheet.append(["delta_hue", self.weights.delta_hue])
            params_sheet.append(["delta_lum", self.weights.delta_lum])
            params_sheet.append(["delta_sat", self.weights.delta_sat])
            params_sheet.append(["delta_edges", self.weights.delta_edges])
            params_sheet.append(["source", video_path])

            ws = wb.create_sheet("Video Analysis", index=vid_index)
            headers = ["Start", "Description", "Image"]
            ws.append(headers)
            header_shift = ws.max_row+1
            
            vid_dir = os.path.join(self.output_dir, f"video_{vid_index}")
            imgs_dir = os.path.join(vid_dir, "images")

            video = open_video(video_path)
            scenes = self.split_video_into_scenes(video, threshold=self.threshold)
            os.makedirs(imgs_dir, exist_ok=True)
            images = scenedetect.save_images(scenes, video, output_dir=imgs_dir, image_name_template='Scene_$SCENE_NUMBER-frame_$FRAME_NUMBER', show_progress=True)

            progress_bar = tqdm(total=len(scenes), desc="Describing Scenes", unit="scene", position=0)
            for i, scene in enumerate(scenes):
                imgs = images[i]
                timestamp = scene[0].get_timecode()
                # frame_index = scene[0].get_frames()

                description = ""
                if self.describe_frames is True:
                    frame = cv2.imread(os.path.join(imgs_dir,imgs[1]))
                    description = self.describe_frame(frame)
                ws.append([timestamp, description])
                cell = ws.cell(i+header_shift, 2)
                alignment = copy.copy(cell.alignment)
                alignment.wrapText = True
                cell.alignment = alignment
                for img_index, img_path in enumerate(imgs):
                    img = ExcelImage(os.path.join(imgs_dir,img_path))
                    img.width = self.img_scale * 16
                    img.height = self.img_scale * 9
                    ws.row_dimensions[i + header_shift].height = img.height * .75
                    col_letter = get_column_letter(3+img_index)
                    ws.add_image(img, "{c}{i}".format(i=i+header_shift,c=col_letter))  # Position image in the correct cell
                    ws.column_dimensions[col_letter].width = self.img_scale * 16 / 7.5
                progress_bar.update(1)

            progress_bar.close()
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 25
            wb.save(os.path.join(vid_dir, "videos_frame_analysis.xlsx"))
            wb.close()
            print("{d}: Done".format(d=datetime.datetime.now().strftime("%Y_%m_%d_%H:%M:%S")))
