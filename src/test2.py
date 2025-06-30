import copy
import datetime

import openpyxl
import scenedetect
import torch
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from transformers import AutoProcessor, AutoModelForCausalLM, AutoConfig
from PIL import Image
import os
from scenedetect import SceneManager, open_video, ContentDetector, FrameTimecode
from tqdm import tqdm
import cv2
import re
from pathlib import Path
import tkinter as tk
from tkinter import filedialog



def get_job_dir(output_dir, desc=None):
    """
    Finds the next job directory index under output_dir.
    Job folders are expected to follow the pattern 'job_<index>'.

    Args:
        output_dir (str): The base directory containing job folders.

    Returns:
        str: Absolute path to the next job directory.
    """
    job_pattern = re.compile(r'^job_(\d+)(_[\w]+)?$')
    max_index = -1

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    for entry in os.listdir(output_dir):
        full_path = os.path.join(output_dir, entry)
        if os.path.isdir(full_path):
            match = job_pattern.match(entry)
            if match:
                idx = int(match.group(1))
                max_index = max(max_index, idx)

    next_index = max_index + 1
    next_job_name = f"job_{next_index:03d}"
    if desc is not None:
        next_job_name += f"_{desc}"
    return os.path.abspath(os.path.join(output_dir, next_job_name))


model_name = "MiaoshouAI/Florence-2-large-PromptGen-v2.0"
processor = None
model = None

device = "cuda:0" if torch.cuda.is_available() else "cpu"
torch_dtype = torch.float16 if torch.cuda.is_available() else torch.float32

def describe_frame(frame, max_tokens=1024):
    global model
    global processor

    if model is None:
        model = AutoModelForCausalLM.from_pretrained(model_name, torch_dtype=torch_dtype, trust_remote_code=True).to(device)
    if processor is None:
        processor = AutoProcessor.from_pretrained(model_name, trust_remote_code=True)
    raw_image = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)).convert('RGB')
    inputs = processor(text="<DETAILED_CAPTION>", images=raw_image, return_tensors="pt").to(device, torch_dtype)
    outputs = model.generate(
        input_ids=inputs["input_ids"],
        pixel_values=inputs["pixel_values"],
        max_new_tokens=max_tokens,
        do_sample=False
    )
    description = processor.batch_decode(outputs, skip_special_tokens=True)[0]
    return description

def select_file():
    root = tk.Tk()
    root.attributes('-topmost', True)
    root.iconify()
    file_path = filedialog.askopenfilenames(
        title="Select a file",
        filetypes=[("All Files", "*.*")]  # You can customize filters
    )
    root.destroy()
    return file_path

path = Path(os.path.abspath(__file__))
ROOT_DIR = path.parent.parent.absolute()
output_dir = "output"
output_dir = os.path.join(ROOT_DIR, output_dir)

DEBUG_MODE = False
describe_frames = True
img_scale = 33
THRESHOLD_DEFAULT = 27
threshold = THRESHOLD_DEFAULT
incrop_ratio=(.4, .2)

video_paths = []

job_desc = input("Job ID: ")
if job_desc == "":
    job_desc = None
threshold = input(f"Threshold({THRESHOLD_DEFAULT}): ")
weights = ContentDetector.Components(delta_edges=.2)
if threshold.strip() == "":
    threshold = THRESHOLD_DEFAULT
else:
    threshold = float(threshold)

job_dir = get_job_dir(output_dir, desc=job_desc)
print("Starting ", os.path.basename(job_dir))
if video_paths is None or len(video_paths) == 0:
    video_paths = select_file()
if len(video_paths) == 0: exit(0)


# params_sheet.append(["Incrop %  (x,y)", str(incrop_ratio)])

def split_video_into_scenes(video, threshold=27.0, weights=ContentDetector.Components()):
    scene_manager = SceneManager()
    scene_manager.add_detector(ContentDetector(threshold=threshold, weights=weights))
    scene_manager.detect_scenes(video, show_progress=True) # , duration=FrameTimecode(2500, video.frame_rate)
    scene_list = scene_manager.get_scene_list()
    return scene_list

for vid_index, video_path in enumerate(video_paths):
    print("{d}: Starting Video {i}".format(i=vid_index+1, d=datetime.datetime.now().strftime("%Y_%m_%d_%H:%M:%S")))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    params_sheet = wb.create_sheet("Parameters")
    params_sheet.append(["Parameter", "Value"])
    params_sheet.append(["threshold", threshold])
    params_sheet.append(["delta_hue", weights.delta_hue])
    params_sheet.append(["delta_lum", weights.delta_lum])
    params_sheet.append(["delta_sat", weights.delta_sat])
    params_sheet.append(["delta_edges", weights.delta_edges])

    ws = wb.create_sheet(f"Frame Scan Video {vid_index}", index=vid_index)
    headers = ["Start", "Description", "Image"]
    ws.append(["Source Video:", video_path])
    ws.append(headers)
    header_shift = ws.max_row+1
    vid_dir = os.path.join(job_dir, f"video_{vid_index}")
    imgs_dir = os.path.join(vid_dir, "images")
    video = open_video(video_path)
    change_frames = {}
    scenes = split_video_into_scenes(video, threshold=threshold)
    os.makedirs(imgs_dir, exist_ok=True)
    images = scenedetect.save_images(scenes, video, output_dir=imgs_dir, image_name_template='Scene_$SCENE_NUMBER-frame_$FRAME_NUMBER', show_progress=True)

    progress_bar = tqdm(total=len(scenes), desc="Describing Scenes", unit="scene", position=0)
    for i, scene in enumerate(scenes):
        imgs = images[i]
        timestamp = scene[0].get_timecode()
        frame_index = scene[0].get_frames()

        description = ""
        if describe_frames is True:
            frame = cv2.imread(os.path.join(imgs_dir,imgs[1]))
            description = describe_frame(frame)
        ws.append([timestamp, description])  # Placeholder for image
        cell = ws.cell(i+header_shift, 2)
        alignment = copy.copy(cell.alignment)
        alignment.wrapText = True
        cell.alignment = alignment
        for img_index, img_path in enumerate(imgs):
            img = ExcelImage(os.path.join(imgs_dir,img_path))
            img.width = img_scale * 16
            img.height = img_scale * 9
            ws.row_dimensions[i + header_shift].height = img.height * .75
            col_letter = get_column_letter(3+img_index)
            ws.add_image(img, "{c}{i}".format(i=i+header_shift,c=col_letter))  # Position image in the correct cell
            ws.column_dimensions[col_letter].width = img_scale * 16 / 7.5
        progress_bar.update(1)

    progress_bar.close()
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    wb.save(os.path.join(vid_dir, "videos_frame_analysis.xlsx"))
    wb.close()
print("{d}: Done".format(d=datetime.datetime.now().strftime("%Y_%m_%d_%H:%M:%S")))
exit(0)

