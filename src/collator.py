import base64
import copy
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk

import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image, ImageTk
import io
import os

MAX_IMAGE_WIDTH = 255

class WorkbookGUI:
    def __init__(self, master, output_dir=None):
        self.master = master
        self.output_dir=output_dir
        self.master.title("Frame Selector Tool")
        self.selected_image_idx = None
        self.focused = False
        self.image_buttons = []
        self.row_data = []  # list of (timestamp, description, images)
        self.current_row_index = 0

        # --- Top Frame: Progress + Exit ---
        top_frame = tk.Frame(master)
        top_frame.pack(pady=5, fill="x")

        top_2_frame = tk.Frame(master)
        top_2_frame.pack(pady=(15,0), fill="x")
        top_3_frame = tk.Frame(master)
        top_3_frame.pack(pady=5, fill="x")

        exit_btn = tk.Button(top_frame, text="Exit", command=lambda: self.quit())
        exit_btn.pack(side="left", padx=5)

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(top_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(side="left", fill="x", expand=True, padx=5)


        finish_btn = tk.Button(top_frame, text="Finish", command=lambda: self.complete())
        finish_btn.pack(side="right", padx=5)

        load_btn = tk.Button(top_frame, text="Load xlsx", command=lambda: self.load_workbook())
        load_btn.pack(side="right", padx=5)



        back_btn = tk.Button(top_2_frame, text="Back (~)", command=lambda: self.prev_row())
        back_btn.pack(side="left", padx=5)
        discard_btn = tk.Button(top_2_frame, text="Discard (`)", command=lambda: self.next_row(False))
        discard_btn.pack(side="right", padx=5)

        keep_btn = tk.Button(top_3_frame, text="Keep (shift+enter)", command=lambda: self.next_row(True))
        keep_btn.pack(side="right", padx=5)


        self.dummy = tk.Frame(self.master)
        self.dummy.pack_forget()  # Hide it

        self.timestamp_label = tk.Label(top_3_frame, text="", font=("Arial", 12))
        self.timestamp_label.pack(pady=5)

        self.description_text = tk.Text(master, height=8, width=80, wrap="word")
        self.description_text.pack(pady=5)

        self.description_text.bind("<FocusIn>", lambda e: self.set_focused(True))
        self.description_text.bind("<FocusOut>", lambda e: self.set_focused(False))

        self.images_frame = tk.Frame(master)
        self.images_frame.pack(pady=5)

        # Key bindings
        master.bind("`", lambda e: self.next_row(False))
        master.bind("<~>", lambda e: self.prev_row())
        master.bind("<Shift-Return>", lambda e: self.next_row(True))
        master.bind("<Tab>", self.toggle_focus)
        self.description_text.bind("<Tab>", self.toggle_focus)
        master.bind("n", lambda e: self.next_row(True))
        for i in range(1, 10):
            master.bind(str(i), self.make_image_selector(i - 1))

        self.load_workbook()

        self.outbook = openpyxl.Workbook()
        self.outsheet = self.outbook.active
        self.outsheet.title = "Collected Frames"
        self.outsheet.append(["Timestamp", "description", "frame base64"])

    def set_focused(self, focused):
        self.focused = focused
        if focused is False:
            self.description_text.config(state='disabled')
        else:
            self.description_text.config(state='normal')
            self.description_text.focus()


    def toggle_focus(self, event=None):
        if not self.focused:
            self.description_text.focus()
        else:
            self.set_focused(False)
            self.dummy.focus_force()
            # self.master.focus()
        # self.focused = not self.focused

    def make_image_selector(self, index):
        def selector(event=None):
            if not self.focused:  # Prevent selection when text is focused
                self.select_image(index)
        return selector

    def select_image(self, index):
        if index < len(self.image_buttons):
            self.selected_image_idx = index
            for i, btn in enumerate(self.image_buttons):
                if i == index:
                    btn.config(highlightbackground="blue", highlightthickness=2)
                else:
                    btn.config(highlightthickness=0)

    def quit(self):
        self.master.destroy()
        exit(0)

    def complete(self):
        file_path = filedialog.asksaveasfilename(initialfile="selected_frames.xlsx", defaultextension=".xlsx", filetypes=[("Excel Workbook", "*.xlsx")])
        self.outbook.save(file_path)
        self.master.destroy()
        exit(0)

    def load_workbook(self):
        file_path = filedialog.askopenfilename(initialdir=self.output_dir, filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            self.master.destroy()
            return
        self.row_data = []
        wb = load_workbook(file_path)
        for sheet in wb.worksheets:
            if sheet.title == "Parameters":
                continue

            image_map = {}
            for img in sheet._images:
                if hasattr(img, 'anchor'):
                    anchor = img.anchor
                    row = anchor._from.row + 1  # Convert to 1-based index
                    image_map.setdefault(row, []).append(img)

            for row in sheet.iter_rows(min_row=2, values_only=False):
                time_cell, desc_cell, *_ = row
                row_index = time_cell.row
                timestamp = time_cell.value
                description = desc_cell.value or ""
                images = image_map.get(row_index, [])
                self.row_data.append([timestamp, description, images])

        if self.row_data:
            self.display_row(0)

    def display_row(self, index):
        if not (0 <= index < len(self.row_data)):
            return
        self.selected_image_idx = None
        self.current_row_index = index
        timestamp, description, images = self.row_data[index]

        self.timestamp_label.config(text=str(timestamp))
        self.description_text.config(state='normal')
        self.description_text.delete("1.0", tk.END)
        self.description_text.insert(tk.END, description)
        self.description_text.config(state='disabled')

        for widget in self.images_frame.winfo_children():
            widget.destroy()
        self.image_buttons.clear()

        for idx, xl_img in enumerate(images):
            try:
                original_ref = xl_img.ref
                xl_img.ref = copy.copy(xl_img.ref)
                image_bytes = xl_img._data()
                xl_img.ref = original_ref
                pil_img = Image.open(io.BytesIO(image_bytes))
                w, h = pil_img.size
                if w > MAX_IMAGE_WIDTH:
                    h = int((MAX_IMAGE_WIDTH / w) * h)
                    w = MAX_IMAGE_WIDTH
                    pil_img = pil_img.resize((w, h), Image.ANTIALIAS)

                tk_img = ImageTk.PhotoImage(pil_img)
                btn = tk.Button(self.images_frame, image=tk_img, borderwidth=2,
                                command=lambda i=idx: self.select_image(i))
                btn.image = tk_img  # Keep reference
                btn.grid(row=0, column=idx, padx=5)
                btn.config(highlightthickness=0)
                self.image_buttons.append(btn)
            except Exception as e:
                print("Failed to process image:", e)

        self.master.lift()
        self.master.focus_force()
        self.select_image(0)
        self.update_progress()

    def update_progress(self):
        total = len(self.row_data)
        if total > 0:
            percent = (self.current_row_index + 1) / total * 100
            self.progress_var.set(percent)

    def next_row(self, keep=False):
        if keep:
            xl_img = self.row_data[self.current_row_index][2][self.selected_image_idx]
            original_ref = xl_img.ref
            xl_img.ref = copy.copy(xl_img.ref)
            image_bytes = xl_img._data()
            xl_img.ref = original_ref
            image_stream = io.BytesIO(image_bytes)
            encoded_img = base64.b64encode(image_stream.getvalue()).decode('utf-8')
            self.row_data[self.current_row_index][1] = self.description_text.get("1.0", "end-1c")
            self.outsheet.append([self.row_data[self.current_row_index][0], self.row_data[self.current_row_index][1], encoded_img])
        if self.current_row_index + 1 < len(self.row_data):
            self.display_row(self.current_row_index + 1)

    def prev_row(self):
        if self.current_row_index == 0 : return
        if self.outsheet.cell(self.outsheet.max_row, 1).value == self.row_data[self.current_row_index - 1][0]:
            self.outsheet.delete_rows(self.outsheet.max_row, 1)
        self.display_row(self.current_row_index - 1)

def main(output_dir=None):
    root = tk.Tk()
    app = WorkbookGUI(root, output_dir=output_dir)
    root.mainloop()

if __name__ == "__main__":
    main()